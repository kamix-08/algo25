import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import os

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
except ImportError:
    raise SystemExit('Brak biblioteki openpyxl. Zainstaluj: pip install openpyxl')

EXCEL_FILE = 'budget.xlsx'
KATEGORIE_PRZYCHODY = ['Wynagrodzenie', 'Premia', 'Freelance', 'Inne przychody']
KATEGORIE_WYDATKI = [
    'Żywność', 'Transport', 'Mieszkanie', 'Rozrywka', 'Zdrowie', 'Odzież', 'Edukacja', 'Inne wydatki'
]
KOLOR_PRZYCHOD = '2ECC71'
KOLOR_WYDATEK = 'E74C3C'
KOLOR_NAGLOWEK = '2C3E50'
KOLOR_WIERSZ_PAR = 'ECF0F1'

def inicjuj_excel(sciezka: str):
    """tworzy plik Excel z nagłówkami, jeśli nie istnieje"""
    if os.path.exists(sciezka):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = 'Transakcje'

    naglowki = ['Data', 'Typ', 'Kategoria', 'Opis', 'Kwota (PLN)']
    ws.append(naglowki)

    fill = PatternFill('solid', fgColor=KOLOR_NAGLOWEK)
    font = Font(bold=True, color='#FFFFFF', size=11)
    border_side = Side(style='thin', color='#FFFFFF')
    border = Border(left = border_side, right = border_side, top = border_side, bottom = border_side)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    szerokosc = [14, 12, 20, 30, 16]
    for i, w in enumerate(szerokosc, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 22
    wb.save(sciezka)

def wczytaj_transakcje(sciezka: str):
    """wczytuje transakcje z arkusza"""
    inicjuj_excel(sciezka)
    wb = load_workbook(sciezka)
    ws = wb.active
    rekordy = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rekordy.append({
                'data': str(row[0]),
                'typ': row[1],
                'kategoria': row[2],
                'opis': row[3],
                'kwota': float(row[4]) if row[4] else 0.0,
            })
    return rekordy

def dodaj_transakcje_excel(sciezka: str, rekord: dict):
    inicjuj_excel(sciezka)
    wb = load_workbook(sciezka)
    ws = wb.active
    nowy_wiersz = ws.max_row + 1

    if nowy_wiersz % 2 == 0:
        bg = PatternFill('solid', fgColor=KOLOR_WIERSZ_PAR)
    else:
        bg = PatternFill('solid', fgColor='FFFFFF')
        
    kolor_kwoty = KOLOR_PRZYCHOD if rekord['typ'] == 'Przychód' else KOLOR_WYDATEK
    
    wartosc = [
        rekord['data'],
        rekord['typ'],
        rekord['kategoria'],
        rekord['opis'],
        rekord['kwota']
    ]
    
    for col, val in enumerate(wartosc, 1):
        cell = ws.cell(row=nowy_wiersz, column=col, value=val)
        cell.fill = bg
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if col == 5:
            cell.number_format = '#,##0.00 PLN'
            cell.font = Font(color=kolor_kwoty, bold=True)
        
    wb.save(sciezka)

def eksportuj_podsumowanie():
    pass

class BudgetManager():
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title('Menedżer Budżetu Osobistego')
        self.root.geometry('900x650')
        self.root.configure(bg='#2C3E50')

        self.excel_sciezka = EXCEL_FILE
        self.buduj_gui()

    def buduj_gui(self):
        naglowek = tk.Frame(self.root, bg='#1A252F', height=60)
        naglowek.pack(fill='x')
        naglowek.pack_propagate(False)
        tk.Label(naglowek, text='Menedżer Budżetu Osobistego', font=('Arial', 20, 'bold'), fg='#FFFFFF', bg='#1A252F').pack(side='left', padx=20, pady=10)

        ramka_plik = tk.Frame(naglowek, bg='#1A252F')
        ramka_plik.pack(side='right', padx=10)
        tk.Label(ramka_plik, text='Plik:', fg='#BDC3C7', bg='#1A252F', font=('Arial', 14)).pack(side='left')
        self.etykieta_plik = tk.Label(ramka_plik, text=self.excel_sciezka, fg='#3498DB', bg='#1A252F', font=('Arial', 14))
        self.etykieta_plik.pack(side='left', padx=4)
        tk.Button(ramka_plik, text='Zmień...', bg='#34495E', fg='#FFFFFF', font=('Arial', 12)).pack(side='left', padx=4)

        glowna = tk.Frame(self.root, bg='#2C3E50')
        glowna.pack(fill='both', expand=True, padx=15, pady=10)

        lewa = tk.Frame(glowna, bg='#2C3E50', width=310)
        lewa.pack(side='left', padx=(0, 10), fill='y')
        lewa.pack_propagate(False)

        prawa = tk.Frame(glowna, bg='#2C3E50')
        prawa.pack(side='left', fill='both', expand=True)

        self.buduj_formularz(lewa)
        self.buduj_saldo(lewa)
        self.buduj_liste(prawa)
        self.buduj_pasek_dolny(prawa)

    def buduj_formularz(self, rodzic):
        ramka = tk.LabelFrame(
            rodzic, text='Nowa transakcja', font=('Arial', 14, 'bold'), fg='#ECF0F1', bg='#34495E', padx=10, pady=10, bd=1, relief='groove')
        ramka.pack(fill='x', padx=5, pady=5)

        pola = [
            ('Typ', 'typ'),
            ('Kategoria', 'kategoria'),
            ('Opis', 'opis'),
            ('Kwota (PLN)', 'kwota'),
            ('Data', 'data')
        ]
        self.zmienne: dict[str, tk.Variable] = {}
        
        for etykieta, klucz in pola:
            tk.Label(ramka, text=etykieta, font=('Arial', 12), fg='#BDC3C7', bg='#34495E', anchor='w').pack(fill='x', padx=10, pady=10)
            
            if klucz == 'typ':
                self.zmienne[klucz] = tk.StringVar(value='Wydatek')
                widget = ttk.Combobox(ramka, textvariable=self.zmienne[klucz], values=['Przychód', 'Wydatek'], state='readonly')
                widget.bind('<<ComboboxSelected>>', lambda e: self.zmiana_typu())
            elif klucz == 'kategoria':
                self.zmienne[klucz] = tk.StringVar(value=KATEGORIE_WYDATKI[0])
                self.cobo_kat = ttk.Combobox(ramka, textvariable=self.zmienne[klucz], values=KATEGORIE_WYDATKI, state='readonly')
                widget = self.cobo_kat
            elif klucz == 'data':
                self.zmienne[klucz] = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
                widget = tk.Entry(ramka, textvariable=self.zmienne[klucz], bg='#ECF0F1', relief='flat', font=('Arial', 12))
            elif klucz == 'kwota':
                self.zmienne[klucz] = tk.StringVar()
                widget = tk.Entry(ramka, textvariable=self.zmienne[klucz], bg='#ECF0F1', relief='flat', font=('Arial', 12))
            else:
                self.zmienne[klucz] = tk.StringVar()
                widget = tk.Entry(ramka, textvariable=self.zmienne[klucz], bg='#ECF0F1', relief='flat', font=('Arial', 12))
                
            widget.pack(fill='x', padx=10, pady=(0,2))
            
        tk.Button(ramka, text='Dodaj i zapisz do Excela', bg='#27AE60', fg='#FFFFFF', font=('Arial', 12), command=self.dodaj_transakcje).pack(padx=10, pady=10)        

    def buduj_saldo(self, rodzic):
        self.ramka_saldo = tk.LabelFrame(rodzic, text='Saldo', font=('Arial', 14, 'bold'), fg='#ECF0F1', bg='#34495E', padx=10, pady=10, bd=1, relief='groove')
        self.ramka_saldo.pack(fill='x', padx=5, pady=5)
        
        self.etykieta_saldo = tk.Label(self.ramka_saldo, text='Ładowanie...', font=('Arial', 16, 'bold'), fg='#27AE60', bg='#34495E')
        self.etykieta_saldo.pack()
        
        self.etykieta_przychody = tk.Label(self.ramka_saldo, text='Przychody: 0 PLN', font=('Arial', 12), fg='#2ECC71', bg='#34495E')
        self.etykieta_przychody.pack()
        
        self.etykieta_wydatki = tk.Label(self.ramka_saldo, text='Wydatki: 0 PLN', font=('Arial', 12), fg='#E74C3C', bg='#34495E')
        self.etykieta_wydatki.pack()

    def buduj_liste(self, rodzic):
        ramka = tk.LabelFrame(rodzic, text='Historia transakcji', font=('Arial', 14, 'bold'), fg='#ECF0F1', bg='#34495E', padx=10, pady=10, bd=1, relief='groove')
        ramka.pack(fill='x', padx=5, pady=5)
        
        kolumny = ('Data', 'Typ', 'Kategoria', 'Opis', 'Kwota (PLN)')
        self.tabela = ttk.Treeview(ramka, columns=kolumny, show='headings', selectmode='extended')
        naglowki = {
            'data': ('Data', 90),
            'typ': ('Typ', 80),
            'kategoria': ('Kategoria', 150),
            'opis': ('Opis', 250),
            'kwota': ('Kwota (PLN)', 120)
        }
        
        for col, (tekst, szer) in naglowki.items():
            self.tabela.heading(col, text=tekst)
            self.tabela.column(col, width=szer, anchor='center')
            
        styl = ttk.Style()
        styl.theme_use('clam')
        styl.configure('Treeview', background='#ECF0F1', foreground='#2C3E50', fieldbackground='#ECF0F1', font=('Arial', 11))
        self.tabela.tag_configure('przychod', foreground='#27AE60')
        self.tabela.tag_configure('wydatek', foreground='#E74C3C')
        
        scroll = ttk.Scrollbar(ramka, orient='vertical', command=self.tabela.yview)
        self.tabela.configure(yscrollcommand=scroll.set)
        scroll.pack(side='right', fill='y')
        self.tabela.pack(fill='both', expand=True)
        
    def buduj_pasek_dolny(self, rodzic):
        pasek = tk.Frame(rodzic, bg='#2C3E50')
        pasek.pack(fill='x', pady=(8, 0))
        przyciski = [
            ('Odśwież', self.odswierz_liste, '#3498DB'),
            ('Eksportuj podsumowanie', self.eksportuj, '#9B59B6'),
            ('Usuń zaznaczone', self.usun_zaznaczone, '#E74C3C'),
        ]
        for tekst, cmd, kolor in przyciski:
            tk.Button(pasek, text=tekst, command=cmd, bg=kolor, fg='#FFFFFF', font=('Arial', 12), padx=12, pady=5).pack(side='left', padx=4)

    def zmiana_typu(self):
        pass

    def dodaj_transakcje(self):
        try:
            kwota = float(self.zmienne['kwota'].get().replace(',', '.'))
        except ValueError:
            messagebox.showerror('Błąd', 'Podana kwota jest nieprawidłowa.')
            return
        
        if kwota <= 0:
            messagebox.showerror('Błąd', 'Kwota musi być większa od zera.')
            return
        
        opis = self.zmienne['opis'].get().strip() or '-'
        rekord = {
            'data': self.zmienne['data'].get(),
            'typ': self.zmienne['typ'].get(),
            'kategoria': self.zmienne['kategoria'].get(),
            'opis': opis,
            'kwota': kwota
        }
        
        try:
            dodaj_transakcje_excel(self.excel_sciezka, rekord)
        except Exception as e:
            messagebox.showerror('Błąd', f'Nie można dodać transakcji: {str(e)}')
            return
        
        self.zmienne['kwota'].set('')
        self.zmienne['opis'].set('')
        self.odswierz_liste()
        messagebox.showinfo('Sukces', f'Transakcja została dodana i zapisana do {os.path.abspath(self.excel_sciezka)}')

    def odswierz_liste(self):
        pass

    def eksportuj(self):
        sciezka = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Plik Excel', '*.xlsx')],
            initialfile='podsumowanie_budzetu.xlsx',
            title='Zapisz podsumowanie'
        )
        if not sciezka:
            return
        try:
            eksportuj_podsumowanie(self.excel_sciezka, sciezka)
            messagebox.showinfo('Eksport zakończony', f'Podsumowanie z wykresem zapisane: \n{sciezka}')
        except Exception as e:
            messagebox.showerror('Błąd eksportu', str(e))

    def usun_zaznaczone(self):
        pass

    def wybierz_plik(self):
        sciezka = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Plik Excel', '*.xlsx')],
            title='Wybierz lub utwórz plik budżetu'
        )
        if sciezka:
            self.excel_sciezka = sciezka
            self.etykieta_plik.config(text=os.path.basename(sciezka))
            
    def

if __name__ == '__main__':
    root = tk.Tk()
    app = BudgetManager(root)
    root.mainloop()